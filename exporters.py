import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)


_HEADER = ["Pokémon", "Carta", "Set", "Rareza", "Precio (€)", "Actualizado"]
_PURPLE = "6C4AB6"
_LIGHT = "EDE7F6"


def _row(card):
    return [
        card["pokemon_name"],
        card["card_name"],
        card["set_name"],
        card["rarity"],
        card["price_current"] if card["price_current"] is not None else "",
        (card["price_updated"] or "")[:10],
    ]


def _write_sheet(ws, cards):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_PURPLE)
    center = Alignment(horizontal="center", vertical="center")

    for col, title in enumerate(_HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    total = 0.0
    for i, card in enumerate(cards, start=2):
        for col, value in enumerate(_row(card), start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = border
            if col == 5:
                cell.number_format = "#,##0.00 €"
                cell.alignment = Alignment(horizontal="right")
        if card["price_current"]:
            total += card["price_current"]

    total_row = len(cards) + 2
    label = ws.cell(row=total_row, column=4, value="TOTAL")
    label.font = Font(bold=True)
    label.alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=5, value=round(total, 2))
    total_cell.font = Font(bold=True)
    total_cell.number_format = "#,##0.00 €"
    total_cell.fill = PatternFill("solid", fgColor=_LIGHT)

    widths = [18, 30, 28, 16, 12, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def build_xlsx(cards):
    """Return an .xlsx file as bytes: separate sheets for Colección and Wishlist."""
    owned = [c for c in cards if c["status"] == "owned"]
    wishlist = [c for c in cards if c["status"] == "wishlist"]

    wb = Workbook()
    ws_owned = wb.active
    ws_owned.title = "Colección"
    _write_sheet(ws_owned, owned)

    ws_wish = wb.create_sheet("Wishlist")
    _write_sheet(ws_wish, wishlist)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _pdf_table(cards, styles):
    data = [_HEADER]
    total = 0.0
    for card in cards:
        price = card["price_current"]
        data.append([
            Paragraph(card["pokemon_name"], styles["cell"]),
            Paragraph(card["card_name"], styles["cell"]),
            Paragraph(card["set_name"], styles["cell"]),
            Paragraph(card["rarity"], styles["cell"]),
            f"{price:.2f}" if price is not None else "-",
            (card["price_updated"] or "")[:10],
        ])
        if price:
            total += price
    data.append(["", "", "", "TOTAL", f"{total:.2f}", ""])

    table = Table(data, colWidths=[28*mm, 42*mm, 38*mm, 26*mm, 22*mm, 24*mm],
                  repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + _PURPLE)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("ALIGN", (5, 0), (5, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),
         [colors.white, colors.HexColor("#F5F2FB")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#" + _LIGHT)),
        ("FONTNAME", (3, -1), (4, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return table


def build_pdf(cards):
    """Return a PDF catalog as bytes with Colección and Wishlist tables."""
    owned = [c for c in cards if c["status"] == "owned"]
    wishlist = [c for c in cards if c["status"] == "wishlist"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=18*mm, bottomMargin=18*mm,
                            leftMargin=14*mm, rightMargin=14*mm,
                            title="Pokédex Tracker")
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle("title", parent=base["Title"],
                                textColor=colors.HexColor("#" + _PURPLE)),
        "h2": ParagraphStyle("h2", parent=base["Heading2"],
                             textColor=colors.HexColor("#" + _PURPLE)),
        "cell": ParagraphStyle("cell", parent=base["BodyText"],
                               fontSize=8, leading=10),
        "meta": ParagraphStyle("meta", parent=base["Normal"],
                               fontSize=9, textColor=colors.grey),
    }

    owned_value = sum(c["price_current"] or 0 for c in owned)
    wish_value = sum(c["price_current"] or 0 for c in wishlist)

    story = [
        Paragraph("Pokédex Tracker", styles["title"]),
        Paragraph(
            f"Generado {datetime.now().strftime('%Y-%m-%d %H:%M')} · "
            f"Colección: {owned_value:.2f} € ({len(owned)} cartas) · "
            f"Wishlist: {wish_value:.2f} € ({len(wishlist)} cartas)",
            styles["meta"]),
        Spacer(1, 8*mm),
        Paragraph("Colección", styles["h2"]),
    ]
    if owned:
        story.append(_pdf_table(owned, styles))
    else:
        story.append(Paragraph("Sin cartas en la colección.", styles["cell"]))
    story.append(Spacer(1, 10*mm))
    story.append(Paragraph("Wishlist", styles["h2"]))
    if wishlist:
        story.append(_pdf_table(wishlist, styles))
    else:
        story.append(Paragraph("Wishlist vacía.", styles["cell"]))

    doc.build(story)
    buf.seek(0)
    return buf.read()
